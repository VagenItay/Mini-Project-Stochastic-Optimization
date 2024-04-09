import math
import os
from openpyxl import load_workbook


def is_greater(point1, point2):
    return point1[0] > point2[0] and point1[1] > point2[1]


def add_arrays(array1, array2):
    result = []
    for val1, val2 in zip(array1, array2):
        result.append(val1 + val2-1)
    return result


def longest_chain_lengths(points):
    # Sort points by x-values and then by y-values
    sorted_points = sorted(points, key=lambda p: (p[0], p[1]))

    n = len(sorted_points)
    lengths = [1] * n  # Initialize lengths for each point

    for i in range(1, n):
        for j in range(i):
            if is_greater(sorted_points[i], sorted_points[j]) and lengths[i] < lengths[j] + 1:
                lengths[i] = lengths[j] + 1

    lengths1 = [1] * n  # Initialize lengths for each point

    for i in range(n, -1, -1):
        for j in range(i, n):
            if is_greater(sorted_points[j], sorted_points[i]) and lengths1[i] < lengths1[j] + 1:
                lengths1[i] = lengths1[j] + 1

    return add_arrays(lengths1, lengths)


def find_weights(longest_chain_lens, maximal_chain):
    return [maximal_chain / element for element in longest_chain_lens]


def find_cube_weights(longest_chain_lens, maximal_chain):
    return [math.sqrt(maximal_chain / element) for element in longest_chain_lens]


def read_points_from_excel(ws):
    # Initialize an empty list to store the points
    points = []
    # Iterate over rows in the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True, max_row=10001, max_col=2):
        # Extract x and y values from the first and second columns respectively
        x = row[0]
        y = row[1]
        if x is not None:
            points.append((x, y))
    return points


def write_weights_into_excel(ws, col, weights):
    for i in range(len(weights)):
        ws.cell(row=i+2, column=col).value = weights[i]


def main():
    # TESTERS: update the input file path
    filename = "data.xlsx"
    file_path = os.path.join(os.path.dirname(__file__), filename)
    wb = load_workbook(filename=file_path, read_only=False)
    i = 1
    for ws in wb:
        points = read_points_from_excel(ws)
        # print(points)

        chain_lengths = longest_chain_lengths(points)
        # print("Lengths of the Longest Chains for Each Point:", chain_lengths)

        longest_chain_length = max(chain_lengths)
        # print("Length of the Longest Chain:", longest_chain_length)

        weights = find_weights(chain_lengths, longest_chain_length)
        write_weights_into_excel(ws, 3, weights)
        # print("Weights for Each Point:", weights)

        print(f"finished {i} iterations.")
        i += 1

        # cubic_weights = find_cube_weights(chain_lengths, longest_chain_length)
        # write_weights_into_excel(ws, 4, cubic_weights)
        # # print("Cubic weights for Each Point:", cubic_weights)
        #
        # print("finished second part.")
        wb.save(filename)
        wb.close()


if __name__ == "__main__":
    main()
